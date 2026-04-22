#!/usr/bin/env python3
"""
SAC 2026 — Extract volunteer data from Excel forms.

Reads the 3 volunteer submission forms (ES, EN, PT) and the selected
volunteers list. Outputs JSON with all volunteer preferences.

Usage:
    cd wca/
    uv run scc-scripts/2026-sac/data/extract_volunteers.py

Input files (in wca/ root):
    - [ES] FORMULARIO VOLUNTARIO SAC 2026 (respuestas).xlsx
    - [EN] VOLUNTEER FORM SAC 2026 (respuestas).xlsx
    - [PT] FORMULÁRIO VOLUNTÁRIO SAC 2026 (respuestas).xlsx
    - SAC2026-registration.xlsx  (definitive approved staff list — Cargo column)

Output:
    scc-scripts/2026-sac/data/volunteers.json
"""

import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: uv pip install openpyxl")
    sys.exit(1)

# Paths
SCRIPT_DIR = Path(__file__).parent
WCA_ROOT = SCRIPT_DIR.parent.parent.parent  # wca/
OUTPUT_FILE = SCRIPT_DIR / "volunteers.json"

FORM_FILES = [
    ("[ES] FORMULARIO VOLUNTARIO SAC 2026 (respuestas).xlsx", "es"),
    ("[EN] VOLUNTEER FORM SAC 2026 (respuestas).xlsx", "en"),
    ("[PT] FORMULÁRIO VOLUNTÁRIO SAC 2026 (respuestas).xlsx", "pt"),
]

SELECTED_FILE = "SAC2026-registration.xlsx"
STAFF_CARGOS = {"Voluntario", "Delegado", "Organizador", "Lider", "Streaming"}


def normalize_wca_id(raw: str) -> str | None:
    """Normalize a WCA ID to uppercase format, or None if invalid."""
    if not raw:
        return None
    raw = raw.strip()
    if raw.upper() in ("N/A", "NA", "NO", "NADA", "NONE", ""):
        return None
    # Check if it looks like a WCA ID (4 digits + 4 letters + 2 digits)
    clean = raw.replace(" ", "")
    if len(clean) == 10 and clean[:4].isdigit() and clean[-2:].isdigit():
        return clean[:4] + clean[4:8].upper() + clean[8:]
    return raw.strip()


def _parse_age(raw) -> int | None:
    """Extract age from various formats like '22', '22.0', '22 años', '17, para cuando...'."""
    if not raw:
        return None
    s = str(raw).strip()
    digits = ""
    for c in s:
        if c.isdigit():
            digits += c
        else:
            break
    return int(digits) if digits else None


def parse_scramble_events(raw: str) -> list[str]:
    """Parse scramble capability string into list of event IDs."""
    if not raw:
        return []
    event_map = {
        "2x2": "222", "3x3": "333", "4x4": "444", "5x5": "555",
        "6x6": "666", "7x7": "777", "clock": "clock",
        "megaminx": "minx", "pyraminx": "pyram", "skewb": "skewb",
        "square": "sq1", "square - 1": "sq1", "square-1": "sq1",
    }
    events = []
    for part in raw.lower().split(","):
        part = part.strip()
        for key, eid in event_map.items():
            if key in part:
                if eid not in events:
                    events.append(eid)
                break
    return events


def extract_forms() -> list[dict]:
    """Extract all volunteer submissions from the 3 forms."""
    all_volunteers = []

    for filename, lang in FORM_FILES:
        filepath = WCA_ROOT / filename
        if not filepath.exists():
            print(f"WARNING: {filename} not found, skipping")
            continue

        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for row in rows:
            if not row[2]:
                continue
            vol = {
                "name": str(row[2]).strip(),
                "email": str(row[1]).strip() if row[1] else "",
                "wca_id": normalize_wca_id(str(row[5]) if row[5] else ""),
                "age": _parse_age(row[3]),
                "language": lang,
                "languages_spoken": str(row[4]).strip() if row[4] else "",
                "preferences": {
                    "main_team": float(row[7]) if row[7] else 0,
                    "data_entry": float(row[8]) if row[8] else 0,
                    "streaming": float(row[9]) if row[9] else 0,
                    "registration": float(row[10]) if row[10] else 0,
                    "wca_booth": float(row[11]) if row[11] else 0,
                    "judge": float(row[12]) if row[12] else 0,
                    "scrambler": float(row[13]) if row[13] else 0,
                    "runner": float(row[14]) if row[14] else 0,
                },
                "can_scramble": parse_scramble_events(str(row[15]) if row[15] else ""),
                "unofficial_comfort": str(row[16]).strip() if row[16] else "",
                "unofficial_scramble": str(row[17]).strip() if row[17] else "",
                "available_days": str(row[18]).strip() if row[18] else "",
                "early_arrival": str(row[19]).strip() if row[19] else "",
                "is_delegate": str(row[20]).strip() if row[20] else "",
                "recommenders": str(row[21]).strip() if row[21] else "",
                "allergies": str(row[22]).strip() if row[22] else "",
                "shirt_size": str(row[23]).strip() if row[23] else "",
                "comments": str(row[24]).strip() if row[24] else "",
            }
            all_volunteers.append(vol)

        wb.close()
        print(f"  {filename}: {len(rows)} submissions")

    return all_volunteers


def extract_selected() -> list[dict]:
    """Extract the approved staff list from SAC2026-registration.xlsx.

    A row counts as approved staff when Status=='a', Registration Status=='accepted',
    and Cargo (column L) is one of STAFF_CARGOS.
    """
    filepath = WCA_ROOT / SELECTED_FILE
    if not filepath.exists():
        print(f"WARNING: {SELECTED_FILE} not found")
        return []

    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    # Sheet 'SAC2026-registration' holds the registration roster
    sheet_name = "SAC2026-registration" if "SAC2026-registration" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Columns: 0=Status 1=Name 2=Country 3=WCA ID 4=Birth Date 5=Gender
    #          6=Email 7=Guests 8=Registration Status 9=Registrant Id
    #          10=Comments(shirt size) 11=Cargo 12=Beneficio
    selected = []
    by_cargo: dict[str, int] = {}
    for row in rows:
        if not row or not row[1]:
            continue
        status = row[0]
        reg_status = row[8] if len(row) > 8 else None
        cargo_raw = row[11] if len(row) > 11 else None
        if status != "a" or reg_status != "accepted" or not cargo_raw:
            continue
        cargo = str(cargo_raw).strip()
        if cargo not in STAFF_CARGOS:
            continue
        selected.append({
            "name": str(row[1]).strip(),
            "country": str(row[2]).strip() if row[2] else "",
            "wca_id": normalize_wca_id(str(row[3]) if row[3] else ""),
            "email": str(row[6]).strip() if row[6] else "",
            "role": cargo,
            "shirt_size": str(row[10]).strip() if len(row) > 10 and row[10] else "",
            "benefit": str(row[12]).strip() if len(row) > 12 and row[12] else "",
        })
        by_cargo[cargo] = by_cargo.get(cargo, 0) + 1

    wb.close()
    print(f"  {SELECTED_FILE} [{sheet_name}]: {len(selected)} approved staff")
    for c, n in sorted(by_cargo.items(), key=lambda x: -x[1]):
        print(f"    {c}: {n}")
    return selected


def match_submissions_to_selected(submissions, selected):
    """Match form submissions to selected volunteers by WCA ID or name."""
    selected_ids = {s["wca_id"] for s in selected if s["wca_id"]}
    selected_names = {s["name"].lower() for s in selected}

    matched = []
    unmatched_submissions = []

    for sub in submissions:
        is_selected = False
        if sub["wca_id"] and sub["wca_id"] in selected_ids:
            is_selected = True
        elif sub["name"].lower() in selected_names:
            is_selected = True

        sub["is_selected"] = is_selected
        if is_selected:
            matched.append(sub)
        else:
            unmatched_submissions.append(sub)

    return matched, unmatched_submissions


def print_summary(submissions, selected, matched):
    """Print summary statistics."""
    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"Total form submissions: {len(submissions)}")
    print(f"Selected volunteers: {len(selected)}")
    print(f"Matched (selected + has form data): {len(matched)}")
    print()

    # Score taker candidates
    score_takers = [v for v in matched if v["preferences"]["data_entry"] >= 8]
    print(f"Score taker candidates (data_entry >= 8): {len(score_takers)}")
    for v in sorted(score_takers, key=lambda x: -x["preferences"]["data_entry"])[:10]:
        print(f"  {v['name']} ({v['wca_id'] or 'N/A'}) - data_entry: {v['preferences']['data_entry']}")

    print()

    # Scramble capabilities
    print("Scramble capabilities (among selected):")
    events_count = {}
    for v in matched:
        for event in v["can_scramble"]:
            events_count[event] = events_count.get(event, 0) + 1
    for event, count in sorted(events_count.items(), key=lambda x: -x[1]):
        print(f"  {event}: {count}")

    print()

    # Role preferences
    print("Average role preferences (among selected):")
    for role in ["judge", "scrambler", "runner", "data_entry"]:
        vals = [v["preferences"][role] for v in matched if v["preferences"][role] > 0]
        avg = sum(vals) / len(vals) if vals else 0
        print(f"  {role}: {avg:.1f}/10")


def main():
    print("SAC 2026 — Volunteer Data Extraction")
    print(f"Working directory: {WCA_ROOT}")
    print()

    print("Reading forms...")
    submissions = extract_forms()

    print("\nReading selected list...")
    selected = extract_selected()

    matched, unmatched = match_submissions_to_selected(submissions, selected)

    print_summary(submissions, selected, matched)

    # Deduplicate by WCA ID (keep first submission)
    seen_ids = set()
    deduped = []
    for v in submissions:
        key = v["wca_id"] or v["email"] or v["name"]
        if key not in seen_ids:
            seen_ids.add(key)
            deduped.append(v)

    # Save output
    output = {
        "metadata": {
            "total_submissions": len(submissions),
            "unique_submissions": len(deduped),
            "selected_count": len(selected),
            "matched_count": len(matched),
        },
        "submissions": deduped,
        "selected": selected,
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"\nSaved to: {OUTPUT_FILE}")
    print(f"  {len(deduped)} unique submissions")
    print(f"  {len(selected)} selected volunteers")


if __name__ == "__main__":
    main()
